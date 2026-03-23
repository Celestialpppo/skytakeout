"""Build template mapping YAML scaffold from workbook headers."""

from __future__ import annotations

from pathlib import Path

import typer
import yaml
from openpyxl import load_workbook

app = typer.Typer()


@app.command()
def main(template_path: Path = typer.Option(...), output_yaml: Path = typer.Option(...)) -> None:
    wb = load_workbook(template_path, data_only=True)
    mappings = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
            if len(row) < 2:
                continue
            left, right = row[0], row[1]
            if left.value and not right.value:
                mappings.append(
                    {
                        "template_id": template_path.stem,
                        "sheet_name": ws.title,
                        "cell": right.coordinate,
                        "field_code": str(left.value).strip(),
                        "required": False,
                        "fill_mode": "overwrite",
                        "postprocess": "identity",
                    }
                )
    output_yaml.parent.mkdir(parents=True, exist_ok=True)
    output_yaml.write_text(yaml.safe_dump({"mappings": mappings}, allow_unicode=True, sort_keys=False), encoding="utf-8")
    typer.echo(f"generated mappings: {len(mappings)}")


if __name__ == "__main__":
    app()
