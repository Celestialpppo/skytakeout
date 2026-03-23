"""XLSX parser implementation."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.models.schemas import Block, BlockType, ParsedDocument
from app.parsers.base import BaseParser


class XlsxParser(BaseParser):
    """Parse xlsx by non-empty cells and simple key-value row pattern."""

    def parse(self, file_path: Path, doc_id: str | None = None) -> ParsedDocument:
        workbook = load_workbook(file_path, data_only=True)
        blocks: list[Block] = []
        idx = 0

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                row_values = [cell.value for cell in row]
                if all(v is None for v in row_values):
                    continue

                if len(row_values) >= 2 and row_values[0] and row_values[1] and all(v is None for v in row_values[2:]):
                    blocks.append(
                        Block(
                            block_id=f"xlsx-{idx}",
                            block_type=BlockType.KV_PAIR,
                            text=f"{row_values[0]}: {row_values[1]}",
                            metadata={"sheet": sheet.title, "row": row[0].row, "key": str(row_values[0]), "value": str(row_values[1])},
                        )
                    )
                    idx += 1

                for cell in row:
                    if cell.value is None:
                        continue
                    blocks.append(
                        Block(
                            block_id=f"xlsx-{idx}",
                            block_type=BlockType.TABLE_CELL,
                            text=str(cell.value).strip(),
                            metadata={"sheet": sheet.title, "row": cell.row, "col": cell.column, "cell": cell.coordinate},
                        )
                    )
                    idx += 1

        return ParsedDocument(
            doc_id=doc_id or file_path.stem,
            source_path=file_path,
            doc_type="xlsx",
            blocks=blocks,
        )
